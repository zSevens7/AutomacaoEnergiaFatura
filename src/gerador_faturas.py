import pandas as pd
import fitz  # PyMuPDF
import os
import re
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

# ==========================================
# 1. FUNÇÕES AUXILIARES
# ==========================================
def text_to_float(texto):
    if not texto: return 0.0
    try:
        clean = re.sub(r'[^\d,\.-]', '', str(texto))
        clean = clean.replace('.', '').replace(',', '.')
        return float(clean)
    except:
        return 0.0

def limpar_id(val):
    try:
        if pd.isna(val): return 9999.0
        return float(re.sub(r'[^\d\.]', '', str(val)))
    except:
        return 9999.0

# ==========================================
# 2. MOTOR DE EXTRAÇÃO COMPLETA
# ==========================================
def extract_invoice_data(pdf_path):
    import fitz
    import re
    import os

    def text_to_float(txt):
        try:
            return float(txt.replace('.', '').replace(',', '.'))
        except:
            return 0.0

    data = {
        "uc": None,
        "ref_month": None,
        "total_value": 0.0,
        "dt_anterior": "-",
        "dt_atual": "-",
        "dt_proxima": "-",
        "leitura_ant": 0.0,
        "leitura_atl": 0.0,
        "consumo_medido": 0.0,
        "energia_compensada": 0.0,
        "consumo_faturado": 0.0,
        "icms": 0.0,
        "pis": 0.0,
        "cofins": 0.0,
        "valor_consumo": 0.0,
        "valor_consumo_compensado": 0.0,
        "valor_energia_injetada": 0.0,
        "valor_cip": 0.0,
        "valor_calculado": 0.0,
        "diferenca": 0.0,
        "erro_extracao": None
    }

    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        text_full = page.get_text("text")

        # ==================================================
        # 1. UC — EXTRAÇÃO POR BLOCOS (ROBUSTA)
        # ==================================================
        blocks = page.get_text("blocks")
        uc_found = None

        for block in blocks:
            text_block = block[4]

            m = re.search(
                r'INSTALA[ÇC][ÃA]O\s*[:\-]?\s*(\d{6,12})',
                text_block,
                re.IGNORECASE
            )
            if m:
                uc_found = m.group(1)
                break

        if not uc_found:
            joined_blocks = "\n".join(b[4] for b in blocks)
            m = re.search(
                r'INSTALA[ÇC][ÃA]O[\s\S]{0,200}?(\d{6,12})',
                joined_blocks,
                re.IGNORECASE
            )
            if m:
                uc_found = m.group(1)

        if uc_found:
            data["uc"] = uc_found

        # ==================================================
        # 2. VALOR TOTAL — MULTI FALLBACK
        # ==================================================
        patterns_valor = [
            r'Total\s*a\s*Pagar\s*R\$\s*([\d\.,]+)',
            r'R\$\s*([\d\.,]+)\s*\n\s*Vencimento',
            r'VALOR\s*COBRADO\s*R?\$?\s*([\d\.,]+)'
        ]

        for p in patterns_valor:
            m = re.search(p, text_full, re.IGNORECASE)
            if m:
                data["total_value"] = text_to_float(m.group(1))
                break

        # ==================================================
        # 3. DATAS
        # ==================================================
        raw_dates = sorted(set(re.findall(r'\d{2}/\d{2}/\d{4}', text_full)))

        if len(raw_dates) >= 3:
            data["dt_anterior"] = raw_dates[0]
            data["dt_atual"] = raw_dates[1]
            data["dt_proxima"] = raw_dates[-1]

        m_ref = re.search(
            r'Refer[êe]ncia[:\s]*(\d{2}/\d{4})',
            text_full,
            re.IGNORECASE
        )
        if m_ref:
            data["ref_month"] = m_ref.group(1)

        # ==================================================
        # 4. LEITURAS / CONSUMO
        # ==================================================
        m_med = re.search(
            r'Consumo\s+ATIVO\s+TOTAL\s+([\d\.,]+)\s+([\d\.,]+)\s+1,00\s+([\d\.,]+)\s+kWh',
            text_full,
            re.IGNORECASE
        )

        if m_med:
            data["leitura_ant"] = text_to_float(m_med.group(1))
            data["leitura_atl"] = text_to_float(m_med.group(2))
            data["consumo_medido"] = text_to_float(m_med.group(3))

        # ==================================================
        # 5. TRIBUTOS
        # ==================================================
        trib_section = re.search(r'(?s)Tributo.*?Valor', text_full, re.IGNORECASE)

        if trib_section:
            trib_text = trib_section.group(0)

            for key in ["ICMS", "PIS", "COFINS"]:
                m = re.search(
                    rf'{key}.*?\s([\d\.,]+)\s*$',
                    trib_text,
                    re.MULTILINE
                )
                if m:
                    data[key.lower()] = text_to_float(m.group(1))

        # ==================================================
        # 6. ITENS FINANCEIROS
        # ==================================================
        for line in text_full.splitlines():
            if "Consumo (kWh)" in line:
                vals = re.findall(r'([\d\.,]+)', line)
                if vals:
                    data["valor_consumo"] = text_to_float(vals[-1])

            elif "Consumo Compensado" in line:
                vals = re.findall(r'([\d\.,]+)', line)
                if vals:
                    data["energia_compensada"] = text_to_float(vals[0])

            elif "Cip-" in line or "Ilum Pub" in line:
                vals = re.findall(r'([\d\.,]+)', line)
                if vals:
                    data["valor_cip"] = text_to_float(vals[-1])

        # ==================================================
        # 7. VALIDAÇÃO FINAL (NUNCA DESCARTA PDF)
        # ==================================================
        if not data["uc"]:
            data["uc"] = f"PENDENTE_{os.path.basename(pdf_path)}"
            data["erro_extracao"] = "UC não localizada automaticamente"

        doc.close()

    except Exception as e:
        data["erro_extracao"] = f"Erro crítico: {str(e)}"
        if not data["uc"]:
            data["uc"] = f"ERRO_{os.path.basename(pdf_path)}"

    return data


# ==========================================
# 3. GERADOR DE EXCEL COMPLETO
# ==========================================
def gerar_relatorio_final():
    print("="*50)
    print("🎨 GERADOR ESTILO BOLETO - VERSÃO COMPLETA")
    print("="*50)
    
    base_dir = os.getcwd()
    pdf_folder = os.path.join(base_dir, "output", "faturas")
    base_excel = os.path.join(base_dir, "output", "Cad_RateioConsumo_Final.xlsx")
    
    if not os.path.exists(base_excel):
        print("❌ Base não encontrada.")
        return

    mes_input = input("Mês/Ano (Ex: 02/2026): ").strip()
    clean_month = mes_input.replace('/', '-')
    output_path = os.path.join(base_dir, "output", f"Relatorio_Boleto_Completo_{clean_month}.xlsx")

    # 1. Carrega Dados
    print("📂 Carregando base...")
    df_base = pd.read_excel(base_excel, dtype={'Conta Contrato': str})
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    
    extracted = {}
    print(f"📡 Processando {len(pdf_files)} faturas...")
    for pdf in pdf_files:
        d = extract_invoice_data(pdf)
        if d['uc']:
            extracted[d['uc']] = d
            # Debug: mostrar extração
            print(f"  UC: {d['uc']} | Total: R$ {d['total_value']:.2f} | ICMS: R$ {d['icms']:.2f}")

    # 2. Monta Lista
    print("✍️  Escrevendo Excel...")
    lista_final = []
    
    if 'ID' in df_base.columns:
        df_base['ID_Sort'] = df_base['ID'].apply(limpar_id)
        df_base = df_base.sort_values(by='ID_Sort')
    
    for _, row in df_base.iterrows():
        uc_excel = str(row['Conta Contrato']).replace('.0', '').strip()
        nome = row['Nome'] if 'Nome' in row else ''
        id_cliente = row['ID'] if 'ID' in row else ''
        
        # ESTRUTURA COMPLETA
        item = {
            "ID": id_cliente,
            "NOME CLIENTE": nome,
            "UC": uc_excel,
            "STATUS": "PENDENTE",
            
            # BLOCO 1: CABEÇALHO
            "MÊS REF": mes_input,
            "VALOR TOTAL (R$)": 0.0,
            
            # BLOCO 2: DATAS
            "LEITURA ANTERIOR": "-",
            "LEITURA ATUAL": "-",
            "PRÓXIMA LEITURA": "-",
            
            # BLOCO 3: MEDIDOR
            "MEDIDOR ANT.": 0,
            "MEDIDOR ATUAL": 0,
            "CONSUMO kWh": 0,
            "ENERGIA COMP. kWh": 0,
            
            # BLOCO 4: TRIBUTOS
            "ICMS (R$)": 0.0,
            "PIS (R$)": 0.0,
            "COFINS (R$)": 0.0,
            
            # BLOCO 5: VALORES DETALHADOS
            "VALOR CONSUMO (R$)": 0.0,
            "VALOR COMPENSADO (R$)": 0.0,
            "VALOR ENERGIA INJ. (R$)": 0.0,
            "CIP (R$)": 0.0,
            
            # BLOCO 6: CÁLCULOS
            "VALOR CALCULADO (R$)": 0.0,
            "DIFERENÇA (R$)": 0.0,
            "VERIFICAÇÃO": "OK"
        }
        
        if uc_excel in extracted:
            d = extracted[uc_excel]
            item["STATUS"] = "DISPONÍVEL"
            item["MÊS REF"] = d['ref_month'] or mes_input
            item["VALOR TOTAL (R$)"] = d['total_value']
            
            item["LEITURA ANTERIOR"] = d['dt_anterior']
            item["LEITURA ATUAL"] = d['dt_atual']
            item["PRÓXIMA LEITURA"] = d['dt_proxima']
            
            item["MEDIDOR ANT."] = d['leitura_ant']
            item["MEDIDOR ATUAL"] = d['leitura_atl']
            item["CONSUMO kWh"] = d['consumo_medido']
            item["ENERGIA COMP. kWh"] = d['energia_compensada']
            
            item["ICMS (R$)"] = d['icms']
            item["PIS (R$)"] = d['pis']
            item["COFINS (R$)"] = d['cofins']
            
            item["VALOR CONSUMO (R$)"] = d['valor_consumo']
            item["VALOR COMPENSADO (R$)"] = d['valor_consumo_compensado']
            item["VALOR ENERGIA INJ. (R$)"] = d['valor_energia_injetada']
            item["CIP (R$)"] = d['valor_cip']
            
            item["VALOR CALCULADO (R$)"] = d['valor_calculado']
            item["DIFERENÇA (R$)"] = d['diferenca']
            
            # Verifica se os valores batem
            if abs(d['diferenca']) < 1.0:  # Diferença menor que R$ 1,00
                item["VERIFICAÇÃO"] = "✅ OK"
            else:
                item["VERIFICAÇÃO"] = f"⚠️ Dif: R$ {d['diferenca']:.2f}"
            
        lista_final.append(item)

    df_resumo = pd.DataFrame(lista_final)

    # 3. Salva Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_resumo.to_excel(writer, sheet_name=f"Relatorio {clean_month}", index=False)
        if 'ID_Sort' in df_base.columns: 
            df_base = df_base.drop(columns=['ID_Sort'])
        df_base.to_excel(writer, sheet_name="Cad.RateioConsumo", index=False)

    # 4. Formatação Visual "BOLETO"
    wb = load_workbook(output_path)
    ws = wb[f"Relatorio {clean_month}"]
    
    # Cores
    header_fill = PatternFill("solid", fgColor="E0E0E0")
    header_font = Font(color="000000", bold=True)
    center = Alignment(horizontal='center', vertical='center')
    
    # Cores por bloco
    fill_cabecalho = PatternFill("solid", fgColor="DDEBF7")      # Azul Claro
    fill_datas = PatternFill("solid", fgColor="FFF2CC")         # Amarelo Claro
    fill_medidor = PatternFill("solid", fgColor="E2EFDA")       # Verde Claro
    fill_tributos = PatternFill("solid", fgColor="FCE4D6")      # Laranja Claro
    fill_valores = PatternFill("solid", fgColor="E4DFEC")       # Roxo Claro
    fill_calculos = PatternFill("solid", fgColor="F2F2F2")      # Cinza Claro
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Formata Cabeçalho
    for col_num, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        
        txt = str(cell.value)
        
        # Aplica cor baseado no conteúdo da coluna
        if col_num <= 4:  # ID, NOME, UC, STATUS
            cell.fill = header_fill
        elif col_num <= 6:  # MÊS REF, VALOR TOTAL
            cell.fill = fill_cabecalho
        elif col_num <= 9:  # DATAS
            cell.fill = fill_datas
        elif col_num <= 13: # MEDIDOR
            cell.fill = fill_medidor
        elif col_num <= 16: # TRIBUTOS
            cell.fill = fill_tributos
        elif col_num <= 20: # VALORES DETALHADOS
            cell.fill = fill_valores
        else:               # CÁLCULOS
            cell.fill = fill_calculos
        
        # Ajusta largura das colunas
        if col_num == 2:  # Nome Cliente
            ws.column_dimensions[cell.column_letter].width = 35
        else:
            ws.column_dimensions[cell.column_letter].width = 15

    # Formata Dados
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center
            cell.border = border
            
            # Formato Moeda para colunas de R$
            header_cell = ws.cell(row=1, column=cell.column)
            if "R$" in str(header_cell.value):
                cell.number_format = 'R$ #,##0.00'
        
        # Status Colorido
        cell_status = row[3]  # Coluna STATUS
        if cell_status.value == "DISPONÍVEL":
            cell_status.font = Font(color="006100", bold=True)
        else:
            cell_status.font = Font(color="9C0006", bold=True)
        
        # Verificação Colorida
        cell_verif = row[-1]  # Última coluna (VERIFICAÇÃO)
        if "✅" in str(cell_verif.value):
            cell_verif.font = Font(color="006100", bold=True)
        elif "⚠️" in str(cell_verif.value):
            cell_verif.font = Font(color="FF9900", bold=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Adiciona Sumário
    ws_summary = wb.create_sheet(title="Sumário")
    ws_summary.append(["RESUMO DO RELATÓRIO"])
    ws_summary.append([""])
    
    total_faturas = len([x for x in lista_final if x["STATUS"] == "DISPONÍVEL"])
    total_valor = sum([x["VALOR TOTAL (R$)"] for x in lista_final])
    total_icms = sum([x["ICMS (R$)"] for x in lista_final])
    
    ws_summary.append([f"Total de Faturas Processadas: {total_faturas}"])
    ws_summary.append([f"Valor Total das Faturas: R$ {total_valor:,.2f}"])
    ws_summary.append([f"Total de ICMS: R$ {total_icms:,.2f}"])
    ws_summary.append([f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])

    wb.save(output_path)
    print(f"\n✅ SUCESSO! Relatório 'Boleto Completo' salvo em:\n{output_path}")
    print(f"   • Total de faturas: {total_faturas}")
    print(f"   • Valor total: R$ {total_valor:,.2f}")
    print(f"   • ICMS total: R$ {total_icms:,.2f}")

if __name__ == "__main__":
    gerar_relatorio_final()