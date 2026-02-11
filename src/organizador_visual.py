import fitz  # PyMuPDF
import os
import re
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter # <--- A CORREÇÃO MÁGICA

# ==========================================
# 1. MOTOR DE EXTRAÇÃO (PyMuPDF)
# ==========================================
def extract_uc_from_pdf(pdf_path):
    """Extrai apenas a UC para conferência rápida"""
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        blocks = page.get_text("blocks")
        blocks.sort(key=lambda b: (b[1], b[0])) 

        for i, b in enumerate(blocks):
            t = b[4].replace('\n', ' ').strip()
            if "Conta Contrato" in t:
                match = re.search(r'(\d{10})', t)
                if match: return match.group(1)
                elif i+1 < len(blocks):
                    match_next = re.search(r'(\d{10})', blocks[i+1][4])
                    if match_next: return match_next.group(1)
    except:
        pass
    return None

# ==========================================
# 2. ORGANIZADOR VISUAL BLINDADO
# ==========================================
def organizar_e_marcar():
    print("="*50)
    print("🎨 ORGANIZADOR VISUAL (VERSÃO BLINDADA)")
    print("="*50)
    
    base_dir = os.getcwd()
    pdf_folder = os.path.join(base_dir, "output", "faturas")
    template_path = os.path.join(base_dir, "data", "Modelo", "202602DemonstrativodeCompensaoAMEL.xlsx")
    
    if not os.path.exists(template_path):
        print("❌ Modelo não encontrado!")
        return

    mes_input = input("Mês/Ano para validar (Ex: 02/2026): ").strip()
    clean_month = mes_input.replace('/', '-')
    output_path = os.path.join(base_dir, "output", f"Demonstrativo_Visual_{clean_month}.xlsx")

    # 1. Mapear PDFs baixados
    print(f"📡 Escaneando PDFs...")
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    ucs_encontradas = set()
    
    for pdf in pdf_files:
        uc = extract_uc_from_pdf(pdf)
        if uc:
            ucs_encontradas.add(uc)
    
    print(f"✅ Faturas identificadas: {len(ucs_encontradas)}")

    # 2. Carregar Excel
    print(f"📂 Abrindo planilha de clientes...")
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        print(f"❌ Erro ao abrir Excel: {e}")
        return
    
    if "Cad.RateioConsumo" not in wb.sheetnames:
        print("❌ Aba 'Cad.RateioConsumo' não existe!")
        return
        
    ws = wb["Cad.RateioConsumo"]

    # 3. Estilos Visuais
    fill_ok = PatternFill("solid", fgColor="C6EFCE")      # Verde
    font_ok = Font(color="006100", bold=True)
    
    fill_missing = PatternFill("solid", fgColor="FFC7CE") # Vermelho
    font_missing = Font(color="9C0006", bold=True)

    # 4. Criar Coluna de Status (BLINDADO)
    status_col_idx = ws.max_column + 1
    col_letter = get_column_letter(status_col_idx) # AA, AB, etc.
    
    # Cabeçalho do Status
    # Vamos escrever na Linha 3 (onde parecem estar os títulos ID, Nome, etc)
    header_row = 3 
    header_cell = ws.cell(row=header_row, column=status_col_idx, value=f"STATUS {mes_input}")
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Ajusta largura
    ws.column_dimensions[col_letter].width = 25 

    # 5. Varrer Clientes e Marcar
    print("\n✍️  Marcando status...")
    count_ok = 0
    count_missing = 0
    
    # IMPORTANTE: Começa da Linha 4 (Pula cabeçalhos e metadados)
    for r in range(4, ws.max_row + 1):
        # Pega a UC da Coluna F (Índice 6)
        cell_uc = ws.cell(row=r, column=6)
        cell_val = cell_uc.value
        
        # Filtro de Segurança: Só processa se tiver valor e parecer um número
        if cell_val:
            uc_excel = str(cell_val).replace('.0', '').strip()
            
            # Se a UC não for numérica (ex: for um texto perdido), ignora
            if not uc_excel.isdigit() or len(uc_excel) < 5:
                continue

            # Célula de destino
            dest_cell = ws.cell(row=r, column=status_col_idx)
            
            if uc_excel in ucs_encontradas:
                dest_cell.value = "DISPONÍVEL"
                dest_cell.fill = fill_ok
                dest_cell.font = font_ok
                count_ok += 1
            else:
                dest_cell.value = "PENDENTE"
                dest_cell.fill = fill_missing
                dest_cell.font = font_missing
                count_missing += 1
                # print(f"   ⚠️ Pendente: UC {uc_excel}") # Descomente se quiser ver lista

    # 6. Salvar
    try:
        wb.save(output_path)
        print(f"\n📊 RESUMO:")
        print(f"   🟩 Encontrados: {count_ok}")
        print(f"   🟥 Pendentes:   {count_missing}")
        print(f"\n💾 Arquivo salvo com sucesso: {output_path}")
    except PermissionError:
        print("❌ ERRO: Feche o arquivo Excel antes de rodar o script!")
    except Exception as e:
        print(f"❌ Erro desconhecido ao salvar: {e}")

if __name__ == "__main__":
    organizar_e_marcar()