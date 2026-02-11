#!/usr/bin/env python3
"""
SISTEMA FINAL PROFISSIONAL - Extração de Faturas Equatorial
Versão definitiva com formatação Excel profissional
"""

import os
import sys
from datetime import datetime
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ==========================================
# CONFIGURAÇÕES DO SISTEMA
# ==========================================
class Config:
    PASTA_FATURAS = "output/faturas"
    BASE_CLIENTES = "output/Cad_RateioConsumo_Final.xlsx"
    PASTA_RELATORIOS = "output/relatorios"
    PASTA_DEBUG = "output/debug"
    
    # Cores para formatação Excel
    CORES = {
        'azul_escuro': "2E75B6",
        'azul_medio': "4472C4",
        'azul_claro': "DDEBF7",
        'verde_escuro': "70AD47",
        'verde_claro': "E2EFDA",
        'laranja_escuro': "ED7D31",
        'laranja_claro': "FCE4D6",
        'amarelo_escuro': "FFC000",
        'amarelo_claro': "FFF2CC",
        'vermelho_escuro': "C00000",
        'vermelho_claro': "FFE6E6",
        'cinza_escuro': "A5A5A5",
        'cinza_medio': "D9D9D9",
        'cinza_claro': "F2F2F2",
        'roxo_claro': "E4DFEC",
    }

# ==========================================
# FUNÇÕES AUXILIARES
# ==========================================
def text_to_float(texto):
    """Converte texto para float de forma robusta"""
    if not texto or texto in ["-", "", " ", "N/A"]:
        return 0.0
    
    try:
        texto = str(texto).strip()
        
        # Remove R$, símbolos, espaços
        texto = re.sub(r'[R\$\s]', '', texto)
        
        # Se for negativo com parênteses
        negativo = False
        if texto.startswith('(') and texto.endswith(')'):
            texto = texto[1:-1]
            negativo = True
        elif texto.startswith('-'):
            texto = texto[1:]
            negativo = True
        
        # Remove caracteres problemáticos no final
        texto = re.sub(r'[.,]+$', '', texto)
        
        # Formato brasileiro (1.234,56) ou (1234,56)
        if ',' in texto:
            partes = texto.split(',')
            if '.' in partes[0]:  # Tem ponto como separador de milhar
                parte_inteira = partes[0].replace('.', '')
            else:
                parte_inteira = partes[0]
            
            if len(partes) == 2:
                parte_decimal = partes[1]  # Aceita todas as casas decimais
                texto = f"{parte_inteira}.{parte_decimal}"
            else:
                texto = parte_inteira
        else:
            # Formato americano ou inteiro
            if texto.count('.') > 1:
                # Remove pontos de milhar (1.234.56 -> 1234.56)
                partes = texto.split('.')
                if len(partes) > 2:
                    inteiro = ''.join(partes[:-1])
                    decimal = partes[-1]
                    texto = f"{inteiro}.{decimal}"
        
        resultado = float(texto)
        return -resultado if negativo else resultado
    
    except Exception as e:
        print(f"⚠️ Conversão: '{texto}' -> 0.0")
        return 0.0

def format_date(date_str):
    """Formata data para dd/mm/yyyy"""
    try:
        if not date_str or date_str in ["-", "", " ", "N/A"]:
            return "-"
        
        date_str = str(date_str).strip()
        
        # Tenta diferentes formatos
        formatos = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%d.%m.%Y']
        for fmt in formatos:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%d/%m/%Y')
            except:
                continue
        
        return date_str
    except:
        return date_str

# ==========================================
# EXTRAÇÃO DE DADOS COMPLETA
# ==========================================
def extrair_dados_fatura(pdf_path):
    """Extrai todos os dados de uma fatura PDF"""
    dados = {
        # Inicializa todas as chaves
        'uc': None, 'instalacao': None, 'ref_month': None,
        'vencimento': None, 'data_emissao': None,
        'dt_anterior': None, 'dt_atual': None, 'dt_proxima': None,
        'leitura_ant': 0, 'leitura_atl': 0, 'consumo_medido': 0,
        'energia_compensada': 0, 'saldo_acumulado': 0,
        'icms': 0.0, 'pis': 0.0, 'cofins': 0.0,
        'icms_aliquota': 0.0, 'pis_aliquota': 0.0, 'cofins_aliquota': 0.0,
        'valor_consumo': 0.0, 'valor_consumo_compensado': 0.0,
        'valor_energia_injetada': 0.0, 'valor_cip': 0.0,
        'valor_adicional_bandeira': 0.0, 'total_value': 0.0,
        'preco_unit_consumo': 0.0, 'preco_unit_compensado': 0.0,
        'tipo_fornecimento': '', 'classificacao': '',
        'bandeira_tarifaria': '', 'cor_bandeira': '',
        'arquivo': os.path.basename(pdf_path),
        'erro_extracao': None
    }
    
    try:
        import fitz
        
        doc = fitz.open(pdf_path)
        pagina = doc[0]
        texto = pagina.get_text("text")
        
        # DEBUG: Salvar texto extraído
        os.makedirs(Config.PASTA_DEBUG, exist_ok=True)
        debug_file = os.path.join(Config.PASTA_DEBUG, f"debug_{os.path.basename(pdf_path)}.txt")
        with open(debug_file, 'w', encoding='utf-8') as f:
            f.write(texto)
        
        # 1. UC (Conta Contrato) - padrões múltiplos
        padroes_uc = [
            r'Conta\s*Contrato\s*(\d{10})',
            r'Contrato\s*(\d{10})',
            r'UC\s*(\d{10})'
        ]
        
        for padrao in padroes_uc:
            match = re.search(padrao, texto, re.IGNORECASE)
            if match:
                dados['uc'] = match.group(1)
                break
        
        # 2. Mês de Referência
        padroes_ref = [
            r'Conta\s*Mês\s*(\d{2}/\d{4})',
            r'REFERÊNCIA\s*(\d{2}/\d{4})',
            r'Referência\s*(\d{2}/\d{4})'
        ]
        
        for padrao in padroes_ref:
            match = re.search(padrao, texto, re.IGNORECASE)
            if match:
                dados['ref_month'] = match.group(1)
                break
        
        # 3. Valor Total
        padroes_total = [
            r'Total\s*a\s*Pagar\s*R\$\s*([\d\.,]+)',
            r'TOTAL\s*A\s*PAGAR\s*R\$\s*([\d\.,]+)',
            r'Valor\s*Documento\s*([\d\.,]+)',
            r'VALOR\s*DOCUMENTO\s*([\d\.,]+)',
            r'R\$\s*([\d\.,]+)\s*Total',
            r'Total\s*R\$\s*([\d\.,]+)'
        ]
        
        for padrao in padroes_total:
            match = re.search(padrao, texto, re.IGNORECASE)
            if match:
                dados['total_value'] = text_to_float(match.group(1))
                break
        
        # 4. Vencimento
        venc_match = re.search(r'Vencimento\s*(\d{2}/\d{2}/\d{4})', texto, re.IGNORECASE)
        if venc_match:
            dados['vencimento'] = format_date(venc_match.group(1))
        
        # 5. Data de Emissão - NOVO
        emissao_patterns = [
            r'Emiss[ãa]o\s*(\d{2}/\d{2}/\d{4})',
            r'Data\s*Emiss[ãa]o\s*(\d{2}/\d{2}/\d{4})',
            r'Data\s*de\s*Emiss[ãa]o\s*(\d{2}/\d{2}/\d{4})',
            r'DATA\s*EMISSÃO\s*(\d{2}/\d{2}/\d{4})',
            r'Emissão:\s*(\d{2}/\d{2}/\d{4})',
            r'Emissão\s*em\s*(\d{2}/\d{2}/\d{4})'
        ]
        
        dados['data_emissao'] = None
        for pattern in emissao_patterns:
            emissao_match = re.search(pattern, texto, re.IGNORECASE)
            if emissao_match:
                dados['data_emissao'] = format_date(emissao_match.group(1))
                break
        
        # Se não encontrou a data de emissão, usar data da leitura atual
        if not dados['data_emissao']:
            # Vamos usar a data atual como fallback, mas será substituída por dt_atual se existir
            dados['data_emissao'] = datetime.now().strftime('%d/%m/%Y')
        
        # 6. Datas de Leitura
        # Padrão: "Leitura Anterior Leitura Atual Nº de Dias Próxima Leitura"
        leitura_pattern = r'Leitura\s*Anterior\s*Leitura\s*Atual.*?(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d+\s+(\d{2}/\d{2}/\d{4})'
        leitura_match = re.search(leitura_pattern, texto, re.IGNORECASE | re.DOTALL)
        
        if leitura_match:
            dados['dt_anterior'] = format_date(leitura_match.group(1))
            dados['dt_atual'] = format_date(leitura_match.group(2))
            dados['dt_proxima'] = format_date(leitura_match.group(3))
            
            # Se não encontrou data de emissão específica, usar data da leitura atual
            if not dados['data_emissao'] or dados['data_emissao'] == datetime.now().strftime('%d/%m/%Y'):
                dados['data_emissao'] = dados['dt_atual']
        
        # 7. Medição
        medicao_pattern = r'(\d+[\.,]\d+)\s+(\d+[\.,]\d+)\s+1,00\s+(\d+[\.,]?\d*)\s+kWh'
        medicao_match = re.search(medicao_pattern, texto)
        if medicao_match:
            dados['leitura_ant'] = text_to_float(medicao_match.group(1))
            dados['leitura_atl'] = text_to_float(medicao_match.group(2))
            dados['consumo_medido'] = text_to_float(medicao_match.group(3))
        
        # 8. Energia Compensada
        compensado_match = re.search(r'Consumo\s*Compensado.*?\(kWh\)\s*(\d+[\.,]\d+)', texto, re.IGNORECASE)
        if compensado_match:
            dados['energia_compensada'] = text_to_float(compensado_match.group(1))
        
        # 9. Saldo Acumulado
        saldo_match = re.search(r'Saldo\s*Acumulado\s*Geral\s*Total:\s*([\d\.,]+)', texto, re.IGNORECASE)
        if saldo_match:
            dados['saldo_acumulado'] = text_to_float(saldo_match.group(1))
        
        # 10. Tributos (valores e alíquotas)
        # Procura tabela de tributos
        tributo_pattern = r'Tributo.*?Base.*?Al[íi]quota.*?Valor.*?(ICMS.*?PIS.*?COFINS.*?)(?=\n\n|\n[A-Z]|\Z)'
        tributo_match = re.search(tributo_pattern, texto, re.IGNORECASE | re.DOTALL)
        
        if tributo_match:
            trib_text = tributo_match.group(1)
            
            # ICMS
            icms_match = re.search(r'ICMS[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', trib_text)
            if icms_match:
                dados['icms'] = text_to_float(icms_match.group(3))
                aliquota = icms_match.group(2).replace('.', '').replace(',', '.')
                try:
                    dados['icms_aliquota'] = float(aliquota) / 100
                except:
                    dados['icms_aliquota'] = 0.0
            
            # PIS
            pis_match = re.search(r'PIS[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', trib_text)
            if pis_match:
                dados['pis'] = text_to_float(pis_match.group(3))
                aliquota = pis_match.group(2).replace('.', '').replace(',', '.')
                try:
                    dados['pis_aliquota'] = float(aliquota) / 100
                except:
                    dados['pis_aliquota'] = 0.0
            
            # COFINS
            cofins_match = re.search(r'COFINS[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', trib_text)
            if cofins_match:
                dados['cofins'] = text_to_float(cofins_match.group(3))
                aliquota = cofins_match.group(2).replace('.', '').replace(',', '.')
                try:
                    dados['cofins_aliquota'] = float(aliquota) / 100
                except:
                    dados['cofins_aliquota'] = 0.0
        
        # 11. Valores Detalhados (Itens de Fatura)
        itens_section = re.search(r'Itens\s*de\s*Fatura.*?(?=ITENS\s*FINANCEIROS|\n\n|\Z)', texto, re.IGNORECASE | re.DOTALL)
        if itens_section:
            itens_text = itens_section.group(0)
            
            # Preço Unitário Consumo
            preco_match = re.search(r'Consumo\s*\(kWh\)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', itens_text)
            if preco_match:
                dados['preco_unit_consumo'] = text_to_float(preco_match.group(2))
        
        # 12. CIP
        cip_match = re.search(r'Cip[^\d]*([\d\.,]+)', texto, re.IGNORECASE)
        if cip_match:
            dados['valor_cip'] = text_to_float(cip_match.group(1))
        
        # 13. Adicional Bandeira (valor)
        bandeira_valor_match = re.search(r'Adicional\s*Bandeira[^\d]*([\d\.,-]+)', texto, re.IGNORECASE)
        if bandeira_valor_match and bandeira_valor_match.group(1).strip():
            dados['valor_adicional_bandeira'] = text_to_float(bandeira_valor_match.group(1))
        
        # 14. Tipo de Fornecimento
        tipo_match = re.search(r'Tipo\s*de\s*Fornecimento:\s*([A-Z]+)', texto, re.IGNORECASE)
        if tipo_match:
            dados['tipo_fornecimento'] = tipo_match.group(1)
        
        # 15. Classificação
        class_match = re.search(r'Classificação:\s*([A-Za-z]+)', texto)
        if class_match:
            dados['classificacao'] = class_match.group(1)
        
        # 16. BANDEIRA TARIFÁRIA (cor: Verde/Amarelo/Vermelho)
        # Padrão: "Band. Tarif.: Verde :" ou "Períodos: Band. Tarif.: Verde"
        bandeira_cor_match = re.search(r'Band\.\s*Tarif\.:\s*([A-Za-z]+)', texto, re.IGNORECASE)
        if bandeira_cor_match:
            bandeira_cor = bandeira_cor_match.group(1).strip().upper()
            dados['bandeira_tarifaria'] = bandeira_cor
            
            # Determina a cor
            if 'VERDE' in bandeira_cor:
                dados['cor_bandeira'] = 'VERDE'
            elif 'AMARELA' in bandeira_cor or 'AMARELO' in bandeira_cor:
                dados['cor_bandeira'] = 'AMARELA'
            elif 'VERMELHA' in bandeira_cor or 'VERMELHO' in bandeira_cor:
                dados['cor_bandeira'] = 'VERMELHA'
            else:
                dados['cor_bandeira'] = bandeira_cor
        
        # 17. Instalação
        instal_match = re.search(r'INSTALAÇÃO:\s*(\d+)', texto, re.IGNORECASE)
        if instal_match:
            dados['instalacao'] = instal_match.group(1)
        
        doc.close()
        
        # Validação
        if not dados['uc']:
            dados['erro_extracao'] = "UC não encontrada"
        elif dados['total_value'] == 0:
            dados['erro_extracao'] = "Valor total não encontrado"
        
        return dados
    
    except Exception as e:
        print(f"❌ Erro no PDF {os.path.basename(pdf_path)}: {e}")
        dados['erro_extracao'] = str(e)
        return dados


# ==========================================
# CÁLCULO DO MÊS DE REFERÊNCIA (REGRA DA PAULA)
# ==========================================
def calcular_mes_competencia(data_leitura_str, dia_corte=12):
    """
    Define a qual relatório a fatura pertence.
    Regra:
    - Leitura > dia 12: Pertence ao mês seguinte.
    - Leitura <= dia 12: Pertence ao mês atual.
    Ex: Leitura 20/01 -> Relatório 02/2026
        Leitura 10/02 -> Relatório 02/2026
    """
    if not data_leitura_str or data_leitura_str == "-":
        return "-"
    
    try:
        data_leitura = datetime.strptime(data_leitura_str, "%d/%m/%Y")
        
        dia = data_leitura.day
        mes = data_leitura.month
        ano = data_leitura.year
        
        # Se leu DEPOIS do dia 12, joga para o próximo mês
        if dia > dia_corte:
            if mes == 12:
                mes = 1
                ano += 1
            else:
                mes += 1
        
        # Retorna no formato MM/AAAA para bater com o que você digita no menu
        return f"{mes:02d}/{ano}"
    
    except:
        return "-"

# ==========================================
# PROCESSAMENTO EM LOTE
# ==========================================
def processar_todas_faturas(mes_referencia):
    """Processa todas as faturas e retorna DataFrame organizado"""
    print("="*70)
    print("📊 PROCESSANDO FATURAS - EQUATORIAL MARANHÃO")
    print("="*70)
    
    if not os.path.exists(Config.PASTA_FATURAS):
        print(f"❌ Pasta não encontrada: {Config.PASTA_FATURAS}")
        return None
    
    arquivos_pdf = glob.glob(os.path.join(Config.PASTA_FATURAS, "*.pdf"))
    if not arquivos_pdf:
        print(f"❌ Nenhum PDF encontrado em: {Config.PASTA_FATURAS}")
        return None
    
    print(f"📁 Pasta: {Config.PASTA_FATURAS}")
    print(f"📅 Mês de Referência: {mes_referencia}")
    print(f"📄 Total de PDFs encontrados: {len(arquivos_pdf)}")
    print("-"*70)
    
    # Carrega base de clientes
    clientes_base = {}
    if os.path.exists(Config.BASE_CLIENTES):
        try:
            df_base = pd.read_excel(Config.BASE_CLIENTES, dtype={'Conta Contrato': str})
            for _, linha in df_base.iterrows():
                uc = str(linha['Conta Contrato']).replace('.0', '').strip()
                nome = linha['Nome'] if 'Nome' in linha else ''
                id_cliente = linha['ID'] if 'ID' in linha else ''
                clientes_base[uc] = {'nome': nome, 'id': id_cliente}
            print(f"✅ Base de clientes carregada: {len(clientes_base)} registros")
        except Exception as e:
            print(f"⚠️ Erro na base: {e}")
    else:
        print("⚠️ Base de clientes não encontrada")
    
    # Processa cada PDF
    resultados = []
    print("\n🔍 EXTRAINDO DADOS:")
    print("-"*50)
    
    for i, pdf_path in enumerate(arquivos_pdf, 1):
        nome_arquivo = os.path.basename(pdf_path)
        print(f"  [{i:3d}/{len(arquivos_pdf):3d}] {nome_arquivo}")
        
        dados = extrair_dados_fatura(pdf_path)
        
        # Define status baseado no erro
        if dados['erro_extracao']:
            dados['status'] = "⚠️ PENDENTE"
            print(f"    ⚠️ Pendência: {dados['erro_extracao']}")
        else:
            dados['status'] = "✅ OK"

        # Garante UC sempre
        if not dados.get('uc'):
            dados['uc'] = f"PENDENTE_{nome_arquivo}"

        # Calcula mês competência se possível
        if dados.get('dt_atual'):
            dados['mes_competencia_calc'] = calcular_mes_competencia(dados['dt_atual'])
        else:
            dados['mes_competencia_calc'] = "-"

        # Dados do cliente
        uc = dados['uc']
        if uc in clientes_base:
            dados['nome_cliente'] = clientes_base[uc]['nome']
            dados['id_cliente'] = clientes_base[uc]['id']
        elif dados['status'] == "✅ OK":
            dados['status'] = "⚠️ SEM BASE"
            dados['nome_cliente'] = "NÃO ENCONTRADO"
            dados['id_cliente'] = ""
        else:
            dados['nome_cliente'] = ""
            dados['id_cliente'] = ""

        resultados.append(dados)

        print(f"    📄 UC: {dados['uc']} | Status: {dados['status']} | Valor: R$ {dados.get('total_value', 0):.2f}")

    
    if not resultados:
        print("\n❌ Nenhuma fatura processada com sucesso")
        return None
    
    print(f"\n{'='*50}")
    print(f"✅ Faturas processadas: {len(resultados)}")
    
    # Organiza dados em DataFrame
    # Na função processar_todas_faturas, substitua a criação do DataFrame por:

    # Organiza dados em DataFrame - VERSÃO CORRIGIDA
    dados_organizados = []
    
    for r in resultados:
        linha = {
            # GRUPO 1: IDENTIFICAÇÃO
            'UC': r.get('uc', ''),
            'INSTALAÇÃO': r.get('instalacao', ''),
            'NOME CLIENTE': r.get('nome_cliente', ''),
            'ID CLIENTE': r.get('id_cliente', ''),
            'STATUS': r.get('status', ''),  # Coluna E é STATUS, não DATA LEITURA
            
            # GRUPO 2: DATAS
            'MÊS REF': r.get('ref_month', ''),
            'MÊS COMPETÊNCIA (CALC)': r.get('mes_competencia_calc', ''),
            'VENCIMENTO': r.get('vencimento', ''),
            'DATA EMISSÃO': r.get('data_emissao', ''),  # Vai pegar a data_emissao se extraída
            'LEITURA ANTERIOR': r.get('dt_anterior', ''),
            'DATA LEITURA': r.get('dt_atual', ''),  # Esta é a data da leitura atual
            'PRÓXIMA LEITURA': r.get('dt_proxima', ''),
            
            # GRUPO 3: MEDIÇÃO (mantém igual)
            'MEDIDOR ANTERIOR (kWh)': r.get('leitura_ant', 0),
            'MEDIDOR ATUAL (kWh)': r.get('leitura_atl', 0),
            'CONSUMO MEDIDO (kWh)': r.get('consumo_medido', 0),
            'ENERGIA COMP. (kWh)': r.get('energia_compensada', 0),
            'SALDO ACUMULADO (kWh)': r.get('saldo_acumulado', 0),
            
            # GRUPO 4: VALORES (R$) (mantém igual)
            'VALOR TOTAL (R$)': r.get('total_value', 0),
            'VALOR CONSUMO (R$)': r.get('valor_consumo', 0),
            'VALOR COMPENSADO (R$)': r.get('valor_consumo_compensado', 0),
            'VALOR ENERGIA INJ. (R$)': r.get('valor_energia_injetada', 0),
            'CIP (R$)': r.get('valor_cip', 0),
            'ADIC. BANDEIRA (R$)': r.get('valor_adicional_bandeira', 0),
            
            # GRUPOS 5-9: (mantém igual)
            'PREÇO UNIT. CONSUMO (R$/kWh)': r.get('preco_unit_consumo', 0),
            'PREÇO UNIT. COMPENSADO (R$/kWh)': r.get('preco_unit_compensado', 0),
            
            'ICMS (R$)': r.get('icms', 0),
            'PIS (R$)': r.get('pis', 0),
            'COFINS (R$)': r.get('cofins', 0),
            
            'ICMS (%)': r.get('icms_aliquota', 0),
            'PIS (%)': r.get('pis_aliquota', 0),
            'COFINS (%)': r.get('cofins_aliquota', 0),
            
            'TIPO FORNECIMENTO': r.get('tipo_fornecimento', ''),
            'CLASSIFICAÇÃO': r.get('classificacao', ''),
            'COR DA BANDEIRA': r.get('cor_bandeira', ''),
            'BANDEIRA TARIF. (INFO)': r.get('bandeira_tarifaria', ''),
            
            'ARQUIVO': r.get('arquivo', ''),
            'ERRO EXTRAÇÃO': r.get('erro_extracao', '')
        }
        dados_organizados.append(linha)
    
    df = pd.DataFrame(dados_organizados)

    # ==========================================================
    # NOVO: FILTRO DA PAULA (DIA 12)
    # Só deixa no Excel o que for do mês digitado no Menu
    # ==========================================================
    if 'MÊS COMPETÊNCIA (CALC)' in df.columns:
        # Filtra o DataFrame
        df_filtrado = df[df['MÊS COMPETÊNCIA (CALC)'] == mes_referencia]
        
        total_arquivos = len(df)
        total_filtrado = len(df_filtrado)
        removidos = total_arquivos - total_filtrado

        if total_filtrado == 0:
            print(f"\n⚠️ AVISO: Nenhuma fatura encontrada para o ciclo {mes_referencia}!")
            print(f"   (Baseado na regra: Dia 13 do mês anterior até dia 12 do mês atual)")
        elif removidos > 0:
            print(f"\n🧹 FILTRO APLICADO: {removidos} faturas de outros meses foram removidas.")
            print(f"   Mantidas apenas as {total_filtrado} faturas do ciclo {mes_referencia}.")
            df = df_filtrado  # Atualiza o DF final apenas com as certas
    # ==========================================================
    
    # Ordena por UC
    if 'UC' in df.columns:
        df = df.sort_values('UC')
        
    return df
    

# ==========================================
# FORMATAÇÃO EXCEL PROFISSIONAL - VERSÃO SIMPLIFICADA E FUNCIONAL
# ==========================================
def formatar_excel_profissional(caminho_excel, df, mes_referencia):
    """Aplica formatação profissional e visual ao Excel"""
    try:
        wb = load_workbook(caminho_excel)
        ws = wb.active
        
        # ==========================================
        # 1. CONFIGURAÇÃO BÁSICA DE ESTILOS
        # ==========================================
        # Cores para grupos de colunas
        cores_grupos = {
            'A': "4F81BD",   # Azul - IDENTIFICAÇÃO
            'F': "F79646",   # Laranja - DATAS
            'M': "9BBB59",   # Verde - MEDIÇÃO
            'R': "C0504D",   # Vermelho - VALORES
            'X': "8064A2",   # Roxo - PREÇOS
            'Z': "4BACC6",   # Azul claro - TRIBUTOS
            'AC': "F2A2C0",  # Rosa - ALÍQUOTAS
            'AF': "948A54",  # Marrom - INFORMAÇÕES
            'AJ': "333333",  # Cinza escuro - ARQUIVO
        }
        
        # Estilos básicos
        fonte_cabecalho = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        fonte_dados = Font(size=9, name="Calibri")
        fonte_titulo = Font(color="1F497D", bold=True, size=14, name="Calibri")
        fonte_subtitulo = Font(color="7F7F7F", italic=True, size=10, name="Calibri")
        
        alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
        alinhamento_direita = Alignment(horizontal="right", vertical="center")
        
        borda_fina = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # ==========================================
        # 2. ADICIONAR TÍTULO
        # ==========================================
        # Insere linhas para título
        ws.insert_rows(1, 2)
        
        # Título principal
        ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
        titulo = ws['A1']
        titulo.value = "⚡ RELATÓRIO DE FATURAS - EQUATORIAL MARANHÃO"
        titulo.font = fonte_titulo
        titulo.fill = PatternFill(start_color="EAF1FF", end_color="EAF1FF", fill_type="solid")
        titulo.alignment = alinhamento_centro
        ws.row_dimensions[1].height = 35
        
        # Subtítulo
        ws.merge_cells(f'A2:{get_column_letter(ws.max_column)}2')
        subtitulo = ws['A2']
        subtitulo.value = f"📅 Mês de Referência: {mes_referencia} | 📊 {len(df)} faturas | ⏰ Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        subtitulo.font = fonte_subtitulo
        subtitulo.alignment = alinhamento_centro
        ws.row_dimensions[2].height = 25
        
        # ==========================================
        # 3. FORMATAR CABEÇALHOS
        # ==========================================
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            celula = ws[f"{col_letter}3"]
            
            # Aplica cor baseada na coluna inicial do grupo
            for inicio_grupo, cor in cores_grupos.items():
                if col_letter >= inicio_grupo:
                    celula.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                    break
            
            # Formatação do texto
            celula.font = fonte_cabecalho
            celula.alignment = alinhamento_centro
            celula.border = borda_fina
        
        # ==========================================
        # 4. FORMATAR DADOS
        # ==========================================
        for linha in range(4, ws.max_row + 1):
            # Linhas zebradas
            if linha % 2 == 0:
                fill_color = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for col in range(1, ws.max_column + 1):
                celula = ws.cell(row=linha, column=col)
                
                # Aplica fundo zebrado (exceto cabeçalho)
                if linha > 3:
                    celula.fill = fill_color
                
                # Aplica bordas
                celula.border = borda_fina
                
                # Formatação baseada no conteúdo
                valor = celula.value
                header = ws.cell(row=3, column=col).value
                
                # Verificação segura do header
                header_str = ""
                if header is not None:
                    header_str = str(header)
                
                # Formatação de MOEDA (R$)
                if header_str and isinstance(header_str, str) and 'R$' in header_str:
                    try:
                        celula.number_format = '"R$" #,##0.0000;[Red]"R$" -#,##0.00000'
                        celula.alignment = alinhamento_direita
                        celula.font = fonte_dados
                    except:
                        pass
                
                # Formatação de PORCENTAGEM (%)
                elif header_str and isinstance(header_str, str) and '%' in header_str:
                    try:
                        celula.number_format = '0.00%'
                        celula.alignment = alinhamento_centro
                        celula.font = fonte_dados
                    except:
                        pass
                
                # Formatação de NÚMEROS (kWh)
                elif header_str and isinstance(header_str, str) and 'kWh' in header_str:
                    try:
                        celula.number_format = '#,##0.00'
                        celula.alignment = alinhamento_direita
                        celula.font = fonte_dados
                    except:
                        pass
                
                # Formatação de DATAS
                elif header_str and isinstance(header_str, str) and any(x in header_str.upper() for x in ['DATA', 'LEITURA', 'VENCIMENTO', 'EMISSÃO']):
                    celula.alignment = alinhamento_centro
                    celula.font = fonte_dados
                
                # Formatação de STATUS
                elif header_str == 'STATUS':
                    celula.alignment = alinhamento_centro
                    if valor == "✅ OK":
                        celula.font = Font(color="00B050", bold=True, name="Calibri")
                        celula.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                    elif "⚠️" in str(valor):
                        celula.font = Font(color="FFC000", bold=True, name="Calibri")
                        celula.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                # Formatação de COR DA BANDEIRA
                elif header_str == 'COR DA BANDEIRA':
                    celula.alignment = alinhamento_centro
                    if valor == 'VERDE':
                        celula.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        celula.font = Font(color="006100", bold=True)
                    elif valor == 'AMARELA':
                        celula.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        celula.font = Font(color="9C6500", bold=True)
                    elif valor == 'VERMELHA':
                        celula.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        celula.font = Font(color="9C0006", bold=True)
                
                # Formatação padrão para texto
                else:
                    celula.alignment = alinhamento_esquerda
                    celula.font = fonte_dados
        
        # ==========================================
        # 5. AJUSTAR LARGURAS DAS COLUNAS
        # ==========================================
        larguras_padrao = {
            'A': 12,   # UC
            'B': 12,   # INSTALAÇÃO
            'C': 30,   # NOME CLIENTE
            'D': 12,   # ID CLIENTE
            'E': 12,   # STATUS
            'F': 10,   # MÊS REF
            'G': 15,   # MÊS COMPETÊNCIA
            'H': 12,   # VENCIMENTO
            'I': 12,   # DATA EMISSÃO
            'J': 12,   # LEITURA ANTERIOR
            'K': 12,   # DATA LEITURA
            'L': 12,   # PRÓXIMA LEITURA
            'M': 15,   # MEDIDOR ANTERIOR
            'N': 15,   # MEDIDOR ATUAL
            'O': 15,   # CONSUMO MEDIDO
            'P': 15,   # ENERGIA COMP.
            'Q': 15,   # SALDO ACUMULADO
            'R': 15,   # VALOR TOTAL
            'S': 15,   # VALOR CONSUMO
            'T': 15,   # VALOR COMPENSADO
            'U': 15,   # VALOR ENERGIA INJ.
            'V': 12,   # CIP
            'W': 15,   # ADIC. BANDEIRA
            'X': 15,   # PREÇO UNIT. CONSUMO
            'Y': 15,   # PREÇO UNIT. COMPENSADO
            'Z': 12,   # ICMS
            'AA': 12,  # PIS
            'AB': 12,  # COFINS
            'AC': 10,  # ICMS %
            'AD': 10,  # PIS %
            'AE': 10,  # COFINS %
            'AF': 15,  # TIPO FORNECIMENTO
            'AG': 15,  # CLASSIFICAÇÃO
            'AH': 15,  # COR DA BANDEIRA
            'AI': 20,  # BANDEIRA TARIF.
            'AJ': 20,  # ARQUIVO
            'AK': 25   # ERRO EXTRAÇÃO
        }
        
        for col_letter, largura in larguras_padrao.items():
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = largura
        
        # ==========================================
        # 6. ADICIONAR RESUMO
        # ==========================================
        ultima_linha = ws.max_row + 3
        
        # Título do resumo
        ws.merge_cells(f'A{ultima_linha}:{get_column_letter(ws.max_column)}{ultima_linha}')
        titulo_resumo = ws.cell(row=ultima_linha, column=1)
        titulo_resumo.value = "📊 RESUMO DO RELATÓRIO"
        titulo_resumo.font = Font(color="1F497D", bold=True, size=12, name="Calibri")
        titulo_resumo.fill = PatternFill(start_color="EAF1FF", end_color="EAF1FF", fill_type="solid")
        titulo_resumo.alignment = alinhamento_centro
        ws.row_dimensions[ultima_linha].height = 25
        
        # Estatísticas
        ultima_linha += 1
        estatisticas = [
            f"Total de Faturas: {len(df)}",
            f"Valor Total: R$ {df['VALOR TOTAL (R$)'].sum():,.2f}",
            f"Consumo Total: {df['CONSUMO MEDIDO (kWh)'].sum():,.0f} kWh",
            f"ICMS Total: R$ {df['ICMS (R$)'].sum():,.2f}",
            f"Média por Fatura: R$ {df['VALOR TOTAL (R$)'].mean():,.2f}"
        ]
        
        col_atual = 1
        colunas_por_item = 6
        
        for estatistica in estatisticas:
            if col_atual <= ws.max_column:
                col_fim = min(col_atual + colunas_por_item - 1, ws.max_column)
                inicio_letra = get_column_letter(col_atual)
                fim_letra = get_column_letter(col_fim)
                
                ws.merge_cells(f'{inicio_letra}{ultima_linha}:{fim_letra}{ultima_linha}')
                celula = ws.cell(row=ultima_linha, column=col_atual)
                celula.value = estatistica
                celula.font = Font(color="2E75B6", bold=True, size=10, name="Calibri")
                celula.alignment = alinhamento_centro
                celula.border = Border(bottom=Side(style='thin', color='2E75B6'))
                
                col_atual = col_fim + 1
        
        # ==========================================
        # 7. CONFIGURAÇÕES FINAIS
        # ==========================================
        # Congelar painéis (cabeçalhos fixos)
        ws.freeze_panes = 'A4'
        
        # Adicionar filtros
        ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row - 5}"
        
        # Ajustar altura das linhas de dados
        for row in range(4, ws.max_row - 5):
            ws.row_dimensions[row].height = 20
        
        # Adicionar assinatura
        ultima_linha = ws.max_row + 2
        ws.merge_cells(f'A{ultima_linha}:{get_column_letter(ws.max_column)}{ultima_linha}')
        assinatura = ws.cell(row=ultima_linha, column=1)
        assinatura.value = f"📋 Gerado automaticamente pelo Sistema de Extração Equatorial - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        assinatura.font = Font(color="7F7F7F", italic=True, size=8, name="Calibri")
        assinatura.alignment = alinhamento_centro
        
        # ==========================================
        # 8. SALVAR
        # ==========================================
        wb.save(caminho_excel)
        print(f"✅ Excel formatado com sucesso!")
        
        return True
    
    except Exception as e:
        print(f"⚠️ Erro na formatação: {e}")
        import traceback
        traceback.print_exc()
        return False
# ==========================================
# FUNÇÃO PARA CRIAR RELATÓRIO FINAL
# ==========================================
def criar_relatorio_final(mes_referencia):
    """Cria relatório final completo"""
    print("\n" + "="*70)
    print("🚀 CRIANDO RELATÓRIO PROFISSIONAL")
    print("="*70)
    
    # Processa faturas
    df = processar_todas_faturas(mes_referencia)
    
    if df is None or df.empty:
        print("❌ Nenhum dado para gerar relatório")
        return None
    
    # Cria pasta de relatórios
    os.makedirs(Config.PASTA_RELATORIOS, exist_ok=True)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"Relatorio_Profissional_{mes_referencia.replace('/', '-')}_{timestamp}.xlsx"
    caminho_completo = os.path.join(Config.PASTA_RELATORIOS, nome_arquivo)
    
    try:
        # Cria Excel com múltiplas abas
        with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
            # Aba 1: DETALHES COMPLETOS (formato bonito)
            df.to_excel(writer, sheet_name='DETALHES COMPLETOS', index=False)
            
            # Aba 2: RESUMO EXECUTIVO
            criar_aba_resumo(writer, df, mes_referencia)
            
            # Aba 3: ESTATÍSTICAS
            criar_aba_estatisticas(writer, df, mes_referencia)
            
            # Aba 4: FATURAS COM ERRO (se houver)
            if 'ERRO EXTRAÇÃO' in df.columns:
                df_erros = df[df['ERRO EXTRAÇÃO'].notna()]
                if not df_erros.empty:
                    df_erros[['ARQUIVO', 'UC', 'ERRO EXTRAÇÃO']].to_excel(
                        writer, sheet_name='ERROS', index=False
                    )
        
        # Aplica formatação profissional na aba principal
        formatar_excel_profissional(caminho_completo, df, mes_referencia)
        
        # Mostrar estatísticas
        mostrar_estatisticas(df, mes_referencia)
        
        print(f"\n{'🎉' * 20}")
        print("🎉 RELATÓRIO PROFISSIONAL CRIADO COM SUCESSO!")
        print(f"{'🎉' * 20}")
        print(f"📍 Local: {caminho_completo}")
        print(f"📊 Total de registros: {len(df)}")
        
        return caminho_completo
    
    except Exception as e:
        print(f"❌ Erro ao criar relatório: {e}")
        return None

def criar_aba_resumo(writer, df, mes_referencia):
    """Cria aba de resumo executivo"""
    # Seleciona colunas importantes para resumo
    colunas_resumo = [
        'UC', 'NOME CLIENTE', 'STATUS', 'MÊS REF', 'VENCIMENTO',
        'VALOR TOTAL (R$)', 'CONSUMO MEDIDO (kWh)', 'ENERGIA COMP. (kWh)',
        'ICMS (R$)', 'CIP (R$)', 'ADIC. BANDEIRA (R$)', 'COR DA BANDEIRA'
    ]
    
    # Filtra colunas existentes
    colunas_existentes = [c for c in colunas_resumo if c in df.columns]
    df_resumo = df[colunas_existentes].copy()
    
    # Ordena por valor total (maiores primeiro)
    if 'VALOR TOTAL (R$)' in df_resumo.columns:
        df_resumo = df_resumo.sort_values('VALOR TOTAL (R$)', ascending=False)
    
    df_resumo.to_excel(writer, sheet_name='RESUMO', index=False)

def criar_aba_estatisticas(writer, df, mes_referencia):
    """Cria aba de estatísticas detalhadas"""
    from openpyxl import Workbook
    
    estatisticas = []
    estatisticas.append(["ESTATÍSTICAS DETALHADAS", ""])
    estatisticas.append(["Mês de Referência:", mes_referencia])
    estatisticas.append(["Data de Geração:", datetime.now().strftime('%d/%m/%Y %H:%M')])
    estatisticas.append(["Total de Faturas:", len(df)])
    estatisticas.append(["", ""])
    
    # Valores monetários
    if 'VALOR TOTAL (R$)' in df.columns:
        total = df['VALOR TOTAL (R$)'].sum()
        media = df['VALOR TOTAL (R$)'].mean()
        maximo = df['VALOR TOTAL (R$)'].max()
        minimo = df['VALOR TOTAL (R$)'].min()
        
        estatisticas.append(["VALORES MONETÁRIOS", ""])
        estatisticas.append(["Valor Total:", f"R$ {total:,.2f}"])
        estatisticas.append(["Valor Médio:", f"R$ {media:,.2f}"])
        estatisticas.append(["Maior Valor:", f"R$ {maximo:,.2f}"])
        estatisticas.append(["Menor Valor:", f"R$ {minimo:,.2f}"])
        estatisticas.append(["", ""])
    
    # Consumo
    if 'CONSUMO MEDIDO (kWh)' in df.columns:
        total = df['CONSUMO MEDIDO (kWh)'].sum()
        media = df['CONSUMO MEDIDO (kWh)'].mean()
        
        estatisticas.append(["CONSUMO DE ENERGIA", ""])
        estatisticas.append(["Consumo Total:", f"{total:,.0f} kWh"])
        estatisticas.append(["Consumo Médio:", f"{media:,.0f} kWh"])
        estatisticas.append(["", ""])
    
    # Tributos
    if 'ICMS (R$)' in df.columns:
        icms_total = df['ICMS (R$)'].sum()
        pis_total = df['PIS (R$)'].sum() if 'PIS (R$)' in df.columns else 0
        cofins_total = df['COFINS (R$)'].sum() if 'COFINS (R$)' in df.columns else 0
        
        estatisticas.append(["TRIBUTOS", ""])
        estatisticas.append(["ICMS Total:", f"R$ {icms_total:,.2f}"])
        estatisticas.append(["PIS Total:", f"R$ {pis_total:,.2f}"])
        estatisticas.append(["COFINS Total:", f"R$ {cofins_total:,.2f}"])
        estatisticas.append(["", ""])
    
    # Bandeiras
    if 'COR DA BANDEIRA' in df.columns:
        estatisticas.append(["DISTRIBUIÇÃO POR BANDEIRA", ""])
        bandeiras = df['COR DA BANDEIRA'].value_counts()
        for bandeira, count in bandeiras.items():
            if bandeira:
                porcentagem = (count / len(df)) * 100
                estatisticas.append([f"  {bandeira}:", f"{count} ({porcentagem:.1f}%)"])
    
    # Status
    if 'STATUS' in df.columns:
        estatisticas.append(["", ""])
        estatisticas.append(["STATUS DAS FATURAS", ""])
        status_counts = df['STATUS'].value_counts()
        for status, count in status_counts.items():
            estatisticas.append([f"  {status}:", f"{count}"])
    
    # Cria DataFrame
    df_stats = pd.DataFrame(estatisticas, columns=["Item", "Valor"])
    df_stats.to_excel(writer, sheet_name='ESTATÍSTICAS', index=False)

def mostrar_estatisticas(df, mes_referencia):
    """Mostra estatísticas no console"""
    print(f"\n📈 ESTATÍSTICAS - {mes_referencia}")
    print("-"*50)
    
    if 'VALOR TOTAL (R$)' in df.columns:
        total = df['VALOR TOTAL (R$)'].sum()
        print(f"💰 Valor Total: R$ {total:,.2f}")
    
    if 'CONSUMO MEDIDO (kWh)' in df.columns:
        consumo = df['CONSUMO MEDIDO (kWh)'].sum()
        print(f"⚡ Consumo Total: {consumo:,.0f} kWh")
    
    if 'ICMS (R$)' in df.columns:
        icms = df['ICMS (R$)'].sum()
        print(f"🏛️  ICMS Total: R$ {icms:,.2f}")
    
    if 'COR DA BANDEIRA' in df.columns:
        print("\n🚦 DISTRIBUIÇÃO POR BANDEIRA:")
        bandeiras = df['COR DA BANDEIRA'].value_counts()
        for bandeira, count in bandeiras.items():
            if bandeira:
                print(f"  • {bandeira}: {count} faturas")
    
    if 'STATUS' in df.columns:
        ok = len(df[df['STATUS'] == '✅ OK'])
        print(f"\n✅ Faturas OK: {ok}/{len(df)}")

# ==========================================
# INTERFACE PRINCIPAL
# ==========================================
def main():
    """Função principal"""
    print("="*70)
    print("⚡ SISTEMA PROFISSIONAL - EXTRATOR DE FATURAS EQUATORIAL")
    print("="*70)
    print("✨ RECURSOS INCLUÍDOS:")
    print("✅ Formatação Excel profissional com grupos coloridos")
    print("✅ Todas as colunas organizadas logicamente")
    print("✅ Cálculo automático do Mês Competência")
    print("✅ Colorização da COR DA BANDEIRA")
    print("✅ Múltiplas abas (Detalhes, Resumo, Estatísticas)")
    print("✅ Legenda automática e estatísticas")
    print("="*70)
    
    while True:
        print("\n" + "="*70)
        print("📋 MENU PRINCIPAL")
        print("="*70)
        print("1. 🎨 Criar relatório profissional (Excel formatado)")
        print("2. 🔍 Testar extração de um arquivo")
        print("3. 📊 Ver estatísticas das pastas")
        print("4. 📖 Ver estrutura do relatório")
        print("5. 🚪 Sair")
        print("-"*70)
        
        try:
            opcao = input("\n👉 Escolha uma opção (1-5): ").strip()
            
            if opcao == '1':
                mes = input("Informe o mês de referência (MM/AAAA): ").strip()
                if not mes:
                    print("⚠️ Mês obrigatório!")
                    continue
                
                print(f"\n⏳ Criando relatório profissional para {mes}...")
                relatorio = criar_relatorio_final(mes)
                
                if relatorio:
                    print(f"\n✅ Relatório criado com sucesso!")
                    print(f"📂 Arquivo: {relatorio}")
                    
                    abrir = input("\n📂 Deseja abrir o arquivo? (s/n): ").strip().lower()
                    if abrir == 's':
                        try:
                            os.startfile(relatorio)
                        except:
                            print("⚠️ Abra manualmente o arquivo")
                
                input("\n⏎ Pressione Enter para continuar...")
            
            elif opcao == '2':
                testar_extracao()
            
            elif opcao == '3':
                mostrar_estatisticas_pastas()
            
            elif opcao == '4':
                mostrar_estrutura_relatorio()
            
            elif opcao == '5':
                print("\n👋 Até logo! Obrigado por usar o sistema.")
                break
            
            else:
                print("❌ Opção inválida!")
        
        except KeyboardInterrupt:
            print("\n\n⚠️ Interrompido pelo usuário")
            break
        except Exception as e:
            print(f"❌ Erro: {e}")

def testar_extracao():
    """Testa extração de um arquivo específico"""
    print("\n🧪 TESTE DE EXTRAÇÃO INDIVIDUAL")
    print("-"*50)
    
    if not os.path.exists(Config.PASTA_FATURAS):
        print(f"❌ Pasta não encontrada: {Config.PASTA_FATURAS}")
        return
    
    arquivos = glob.glob(os.path.join(Config.PASTA_FATURAS, "*.pdf"))
    if not arquivos:
        print("❌ Nenhum PDF encontrado")
        return
    
    print("📄 Arquivos disponíveis:")
    for i, arq in enumerate(arquivos[:5], 1):
        print(f"  {i}. {os.path.basename(arq)}")
    
    try:
        escolha = int(input(f"\nEscolha (1-{min(5, len(arquivos))}): "))
        if 1 <= escolha <= len(arquivos):
            arquivo = arquivos[escolha-1]
            print(f"\n🔍 Testando: {os.path.basename(arquivo)}")
            
            dados = extrair_dados_fatura(arquivo)
            if dados and dados.get('uc'):
                print("\n📋 DADOS EXTRAÍDOS:")
                print("-"*40)
                
                # Agrupa por categoria
                categorias = {
                    'IDENTIFICAÇÃO': ['uc', 'instalacao', 'ref_month', 'vencimento', 'data_emissao'],
                    'DATAS LEITURA': ['dt_anterior', 'dt_atual', 'dt_proxima'],
                    'MEDIÇÃO': ['leitura_ant', 'leitura_atl', 'consumo_medido'],
                    'ENERGIA GD': ['energia_compensada', 'saldo_acumulado'],
                    'VALORES': ['total_value', 'valor_cip', 'valor_adicional_bandeira'],
                    'PREÇOS': ['preco_unit_consumo', 'preco_unit_compensado'],
                    'TRIBUTOS': ['icms', 'pis', 'cofins'],
                    'INFORMAÇÕES': ['tipo_fornecimento', 'classificacao', 'bandeira_tarifaria', 'cor_bandeira']
                }
                
                for cat, campos in categorias.items():
                    print(f"\n{cat}:")
                    for campo in campos:
                        if campo in dados and dados[campo]:
                            valor = dados[campo]
                            if isinstance(valor, float):
                                if campo.endswith('_aliquota'):
                                    print(f"  {campo:30}: {valor:.2%}")
                                elif campo in ['total_value', 'valor_cip', 'valor_adicional_bandeira', 
                                             'icms', 'pis', 'cofins', 'preco_unit_consumo', 'preco_unit_compensado']:
                                    print(f"  {campo:30}: R$ {valor:,.5f}")
                                else:
                                    print(f"  {campo:30}: {valor:,.3f}")
                            else:
                                print(f"  {campo:30}: {valor}")
                
                if dados.get('erro_extracao'):
                    print(f"\n⚠️ ERRO: {dados['erro_extracao']}")
                
                # Calcula mês competência
                if dados.get('dt_atual'):
                    mes_comp = calcular_mes_competencia(dados['dt_atual'])
                    print(f"\n📅 MÊS COMPETÊNCIA (calc): {mes_comp}")
            else:
                print("❌ Falha na extração")
        else:
            print("❌ Escolha inválida")
    except ValueError:
        print("❌ Digite um número")
    
    input("\n⏎ Pressione Enter para continuar...")

def mostrar_estatisticas_pastas():
    """Mostra estatísticas das pastas"""
    print("\n📁 ESTATÍSTICAS DAS PASTAS")
    print("="*50)
    
    pastas = [
        (Config.PASTA_FATURAS, "📄 Faturas PDF"),
        (Config.PASTA_RELATORIOS, "📊 Relatórios"),
        (Config.PASTA_DEBUG, "🔧 Debug"),
    ]
    
    for caminho, nome in pastas:
        if os.path.exists(caminho):
            itens = len(glob.glob(os.path.join(caminho, "*")))
            tamanho = 0
            for arq in glob.glob(os.path.join(caminho, "*")):
                try:
                    tamanho += os.path.getsize(arq)
                except:
                    pass
            
            print(f"\n{nome}:")
            print(f"  📍 {caminho}")
            print(f"  📦 Itens: {itens}")
            print(f"  💾 Tamanho: {tamanho/1024/1024:.1f} MB")
        else:
            print(f"\n{nome}: ❌ NÃO EXISTE")
    
    print(f"\n📄 Base de clientes:")
    if os.path.exists(Config.BASE_CLIENTES):
        print(f"  ✅ {Config.BASE_CLIENTES}")
    else:
        print(f"  ❌ {Config.BASE_CLIENTES} (NÃO ENCONTRADA)")
    
    input("\n⏎ Pressione Enter para continuar...")

def mostrar_estrutura_relatorio():
    """Mostra estrutura do relatório"""
    print("\n📋 ESTRUTURA DO RELATÓRIO PROFISSIONAL")
    print("="*60)
    
    grupos = [
        ("🔷 IDENTIFICAÇÃO (5 colunas)", [
            "• UC", "• INSTALAÇÃO", "• NOME CLIENTE", 
            "• ID CLIENTE", "• STATUS"
        ]),
        
        ("📅 DATAS (7 colunas)", [
            "• MÊS REF", "• MÊS COMPETÊNCIA (CALC)", 
            "• VENCIMENTO", "• DATA EMISSÃO",
            "• LEITURA ANTERIOR", "• DATA LEITURA", 
            "• PRÓXIMA LEITURA"
        ]),
        
        ("⚡ MEDIÇÃO (5 colunas)", [
            "• MEDIDOR ANTERIOR (kWh)", "• MEDIDOR ATUAL (kWh)",
            "• CONSUMO MEDIDO (kWh)", "• ENERGIA COMP. (kWh)",
            "• SALDO ACUMULADO (kWh)"
        ]),
        
        ("💰 VALORES (R$) (6 colunas)", [
            "• VALOR TOTAL (R$)", "• VALOR CONSUMO (R$)",
            "• VALOR COMPENSADO (R$)", "• VALOR ENERGIA INJ. (R$)",
            "• CIP (R$)", "• ADIC. BANDEIRA (R$)"
        ]),
        
        ("📊 PREÇOS (2 colunas)", [
            "• PREÇO UNIT. CONSUMO (R$/kWh)",
            "• PREÇO UNIT. COMPENSADO (R$/kWh)"
        ]),
        
        ("🏛️ TRIBUTOS (R$) (3 colunas)", [
            "• ICMS (R$)", "• PIS (R$)", "• COFINS (R$)"
        ]),
        
        ("📈 ALÍQUOTAS (%) (3 colunas)", [
            "• ICMS (%)", "• PIS (%)", "• COFINS (%)"
        ]),
        
        ("ℹ️ INFORMAÇÕES (4 colunas)", [
            "• TIPO FORNECIMENTO", "• CLASSIFICAÇÃO",
            "• COR DA BANDEIRA (colorida)", "• BANDEIRA TARIF. (INFO)"
        ]),
        
        ("📁 ARQUIVO (2 colunas)", [
            "• ARQUIVO", "• ERRO EXTRAÇÃO"
        ])
    ]
    
    for titulo, colunas in grupos:
        print(f"\n{titulo}")
        for coluna in colunas:
            print(f"  {coluna}")
    
    print("\n🎨 FORMATAÇÃO INCLUÍDA:")
    print("  • Cabeçalhos coloridos por grupo")
    print("  • Células da bandeira coloridas automaticamente")
    print("  • Formatação de moeda, porcentagem, números")
    print("  • Filtros automáticos em todas as colunas")
    print("  • Título, legenda e estatísticas")
    print("  • Múltiplas abas (Detalhes, Resumo, Estatísticas)")
    
    input("\n⏎ Pressione Enter para continuar...")

# ==========================================
# EXECUÇÃO
# ==========================================
if __name__ == "__main__":
    try:
        # Cria pastas necessárias
        for pasta in [Config.PASTA_FATURAS, Config.PASTA_RELATORIOS, Config.PASTA_DEBUG]:
            os.makedirs(pasta, exist_ok=True)
        
        main()
    except KeyboardInterrupt:
        print("\n\n👋 Programa interrompido")
    except Exception as e:
        print(f"\n❌ ERRO CRÍTICO: {e}")
        import traceback
        traceback.print_exc()