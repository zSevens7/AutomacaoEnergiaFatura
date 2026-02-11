import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


# ==========================
# Utils
# ==========================
def normalizar_acesso_equatorial(valor):
    """
    Se for email -> mantém
    Se for data -> converte para DD/MM/AAAA
    Tudo retorna STRING
    """
    if pd.isna(valor):
        return ""

    valor_str = str(valor).strip()

    # Email
    if "@" in valor_str:
        return valor_str.lower()

    # Tentar interpretar como data
    try:
        data = pd.to_datetime(valor, errors="coerce", dayfirst=True)
        if pd.notna(data):
            return data.strftime("%d/%m/%Y")
    except Exception:
        pass

    return valor_str

def corrigir_cpf_cnpj(valor):
    """
    Remove não numéricos e adiciona zeros à esquerda.
    - Se len <= 11: Padroniza para 11 (CPF)
    - Se len > 11: Padroniza para 14 (CNPJ)
    """
    if pd.isna(valor):
        return ""
    
    # Deixa apenas números
    limpo = re.sub(r"\D", "", str(valor))
    
    if not limpo:
        return ""

    # Lógica do Zero à Esquerda
    if len(limpo) <= 11:
        return limpo.zfill(11) # Ex: 6483747349 -> 06483747349
    else:
        return limpo.zfill(14) # Garante CNPJ correto também

# ==========================
# Main
# ==========================
def gerar_planilha_rateio():
    base_dir = os.getcwd()

    caminho_entrada = os.path.join(
        base_dir,
        "data",
        "Modelo",
        "202602DemonstrativodeCompensaoAMEL.xlsx"
    )

    pasta_output = os.path.join(base_dir, "output")
    os.makedirs(pasta_output, exist_ok=True)

    # --------------------------
    # Leitura da base original
    # --------------------------
    df = pd.read_excel(
        caminho_entrada,
        sheet_name="Cad.RateioConsumo",
        header=2
    )

    df.columns = df.columns.astype(str).str.strip()

    # Remove linhas inválidas
    df = df.dropna(subset=["CNPJ/CPF", "Conta Contrato"])

    # --------------------------
    # Limpezas
    # --------------------------

    # --- CORREÇÃO APLICADA AQUI ---
    # Aplica a função que coloca o zero à esquerda
    df["CNPJ/CPF"] = df["CNPJ/CPF"].apply(corrigir_cpf_cnpj)

    # Corrigir coluna M (email OU data nascimento)
    if "Acesso equatorial" in df.columns:
        df["Acesso equatorial"] = df["Acesso equatorial"].apply(
            normalizar_acesso_equatorial
        )

    # Data de Inicio (se existir)
    if "Data de Inicio" in df.columns:
        df["Data de Inicio"] = (
            pd.to_datetime(
                df["Data de Inicio"],
                errors="coerce",
                dayfirst=True
            )
            .dt.strftime("%d/%m/%Y")
            .astype(str)
        )

    # --------------------------
    # Criação do Excel FINAL
    # --------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Cad.RateioConsumo"

    COLUNAS_MODELO = [
        "Nome",
        "ID",
        "CNPJ/CPF",
        "Conta Contrato",
        "Tipo Instalação",
        "Consumo Médio (kWh)",
        "Desconto",
        "Desconto Band. Tar.",
        "Usina Associada",
        "Percentual Rateio (%)",
        "Acesso equatorial",
        "Endereço (igual a conta)",
        "Razão do Titular",
        "Data de Inicio",
        "Vigente"
    ]

    # Estilos
    header_fill = PatternFill("solid", fgColor="9BBB59")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    # Header
    for col, nome in enumerate(COLUNAS_MODELO, start=1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = align
        ws.column_dimensions[chr(64 + col)].width = 22

    # Dados (FORÇANDO TEXTO PARA NÃO PERDER O ZERO NO EXCEL FINAL)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for col, nome in enumerate(COLUNAS_MODELO, start=1):
            valor = str(row[nome]) if nome in df.columns else ""
            cell = ws.cell(row=i, column=col, value=valor)
            cell.number_format = "@" # Formato Texto

    ws.freeze_panes = "A2"

    caminho_saida = os.path.join(
        pasta_output,
        "Cad_RateioConsumo_Final.xlsx"
    )

    wb.save(caminho_saida)
    print("✅ Planilha gerada corretamente (CPFs corrigidos com zero à esquerda)")


if __name__ == "__main__":
    gerar_planilha_rateio()