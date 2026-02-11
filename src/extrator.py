import pandas as pd
import fitz  # PyMuPDF
import os
import re
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import json

# ==========================================
# 1. FUNÇÕES AUXILIARES
# ==========================================
def text_to_float(texto):
    if not texto: return 0.0
    try:
        clean = re.sub(r'[^\d,\.-]', '', str(texto))
        clean = clean.replace('.', '').replace(',', '.')
        return float(clean)
    except:
        return 0.0

def limpar_id(val):
    try:
        if pd.isna(val): return 9999.0
        return float(re.sub(r'[^\d\.]', '', str(val)))
    except:
        return 9999.0

def format_date(date_str):
    """Formata data para dd/mm/yyyy"""
    try:
        if not date_str or date_str == "-":
            return "-"
        # Remove espaços e formata
        date_str = str(date_str).strip()
        # Tenta diferentes formatos
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%d/%m/%Y')
            except:
                continue
        return date_str
    except:
        return date_str

# ==========================================
# 2. MOTOR DE EXTRAÇÃO COMPLETO
# ==========================================
def extract_invoice_data(pdf_path):
    """
    Extrai TODOS os dados possíveis da fatura PDF da Equatorial
    """
    data = {
        # DADOS BÁSICOS
        "uc": None,
        "ref_month": None,
        "total_value": 0.0,
        "vencimento": None,
        "data_emissao": None,
        
        # DATAS DE LEITURA
        "dt_anterior": "-",
        "dt_atual": "-", 
        "dt_proxima": "-",
        
        # MEDIÇÃO
        "leitura_ant": 0,
        "leitura_atl": 0,
        "consumo_medido": 0,
        
        # ENERGIA GD
        "energia_compensada": 0,
        "consumo_faturado": 0,
        "saldo_acumulado": 0,
        
        # TRIBUTOS (VALORES)
        "icms": 0.0,
        "pis": 0.0, 
        "cofins": 0.0,
        
        # TRIBUTOS (ALÍQUOTAS)
        "icms_aliquota": 0.0,
        "pis_aliquota": 0.0,
        "cofins_aliquota": 0.0,
        
        # VALORES DETALHADOS
        "valor_consumo": 0.0,
        "valor_consumo_compensado": 0.0,
        "valor_energia_injetada": 0.0,
        "valor_adicional_bandeira": 0.0,
        "valor_cip": 0.0,
        
        # PREÇOS UNITÁRIOS
        "preco_unit_consumo": 0.0,
        "preco_unit_compensado": 0.0,
        
        # CÁLCULOS
        "valor_calculado": 0.0,
        "diferenca": 0.0,
        
        # INFORMAÇÕES ADICIONAIS
        "tipo_fornecimento": "",
        "classificacao": "",
        "cnpj": "",
        "instalacao": "",
        "bandeira_tarifaria": "",
        
        # STATUS
        "processado": False,
        "erro": None
    }

    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        text_full = page.get_text("text") 
        
        # DEBUG: Salvar texto extraído para análise
        # with open(f"debug_{os.path.basename(pdf_path)}.txt", "w", encoding="utf-8") as f:
        #     f.write(text_full)
        
        # --- A. DADOS BÁSICOS ---
        
        # UC (Conta Contrato)
        uc_match = re.search(r'Conta\s*Contrato\s*(\d{10})', text_full, re.IGNORECASE)
        if not uc_match:
            uc_match = re.search(r'Contrato\s*(\d{10})', text_full, re.IGNORECASE)
        if uc_match:
            data['uc'] = uc_match.group(1)
        
        # Mês de Referência
        ref_match = re.search(r'Conta\s*Mês\s*(\d{2}/\d{4})', text_full, re.IGNORECASE)
        if not ref_match:
            ref_match = re.search(r'REFERÊNCIA\s*(\d{2}/\d{4})', text_full, re.IGNORECASE)
        if ref_match:
            data['ref_month'] = ref_match.group(1)
        
        # Valor Total
        total_match = re.search(r'Total\s*a\s*Pagar\s*R\$\s*([\d\.,]+)', text_full, re.IGNORECASE)
        if not total_match:
            total_match = re.search(r'VALOR\s*DOCUMENTO\s*([\d\.,]+)', text_full, re.IGNORECASE)
        if total_match:
            data['total_value'] = text_to_float(total_match.group(1))
        
        # Data de Vencimento
        venc_match = re.search(r'Vencimento\s*(\d{2}/\d{2}/\d{4})', text_full, re.IGNORECASE)
        if not venc_match:
            venc_match = re.search(r'VENCIMENTO\s*(\d{2}/\d{2}/\d{4})', text_full, re.IGNORECASE)
        if venc_match:
            data['vencimento'] = format_date(venc_match.group(1))
        
        # Data de Emissão
        emissao_match = re.search(r'DATA\s*DE\s*EMISSÃO:\s*(\d{2}/\d{2}/\d{4})', text_full, re.IGNORECASE)
        if emissao_match:
            data['data_emissao'] = format_date(emissao_match.group(1))
        
        # --- B. DATAS DE LEITURA ---
        dates = re.findall(r'(\d{2}/\d{2}/\d{4})', text_full)
        # Filtrar datas que parecem ser de leitura (evitar datas muito antigas)
        current_year = datetime.now().year
        valid_dates = []
        
        for date_str in dates:
            try:
                dt = datetime.strptime(date_str, '%d/%m/%Y')
                # Considerar apenas datas dos últimos 2 anos
                if dt.year >= current_year - 1:
                    valid_dates.append(date_str)
            except:
                continue
        
        if len(valid_dates) >= 3:
            # Ordenar datas
            valid_dates = sorted(set(valid_dates))
            data['dt_anterior'] = format_date(valid_dates[0])
            data['dt_atual'] = format_date(valid_dates[1])
            data['dt_proxima'] = format_date(valid_dates[-1])
        
        # --- C. MEDIÇÃO ---
        # Padrão: número, número, 1,00, número kWh
        consumo_match = re.search(r'(\d+[\.,]\d+)\s+(\d+[\.,]\d+)\s+1,00\s+(\d+[\.,]?\d*)\s+kWh', text_full)
        if consumo_match:
            data['leitura_ant'] = text_to_float(consumo_match.group(1))
            data['leitura_atl'] = text_to_float(consumo_match.group(2))
            data['consumo_medido'] = text_to_float(consumo_match.group(3))
            data['consumo_faturado'] = data['consumo_medido']
        
        # --- D. ENERGIA GD ---
        # Consumo Compensado
        compensado_match = re.search(r'CONSUMO\s*COMPENSADO.*?\((\d+[\.,]\d+)\s*kWh\)', text_full, re.IGNORECASE | re.DOTALL)
        if not compensado_match:
            compensado_match = re.search(r'Consumo\s*Compensado.*?\(kWh\)\s*(\d+[\.,]\d+)', text_full, re.IGNORECASE)
        if compensado_match:
            data['energia_compensada'] = text_to_float(compensado_match.group(1))
        
        # Saldo Acumulado Geral Total
        saldo_match = re.search(r'Saldo\s*Acumulado\s*Geral\s*Total:\s*([\d\.,]+)', text_full, re.IGNORECASE)
        if saldo_match:
            data['saldo_acumulado'] = text_to_float(saldo_match.group(1))
        
        # --- E. TRIBUTOS - VALORES E ALÍQUOTAS ---
        # Procura pela tabela de tributos
        tributo_pattern = r'Tributo.*?Base.*?Al[íi]quota.*?Valor.*?(ICMS.*?PIS.*?COFINS.*?)(?=\n\n|\n[A-Z]|\Z)'
        tributo_match = re.search(tributo_pattern, text_full, re.IGNORECASE | re.DOTALL)
        
        if tributo_match:
            tributo_text = tributo_match.group(1)
            
            # ICMS
            icms_match = re.search(r'ICMS[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', tributo_text)
            if icms_match:
                data['icms'] = text_to_float(icms_match.group(3))  # Valor do ICMS
                data['icms_aliquota'] = text_to_float(icms_match.group(2))  # Alíquota do ICMS
            
            # PIS
            pis_match = re.search(r'PIS[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', tributo_text)
            if pis_match:
                data['pis'] = text_to_float(pis_match.group(3))  # Valor do PIS
                data['pis_aliquota'] = text_to_float(pis_match.group(2))  # Alíquota do PIS
            
            # COFINS
            cofins_match = re.search(r'COFINS[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)', tributo_text)
            if cofins_match:
                data['cofins'] = text_to_float(cofins_match.group(3))  # Valor do COFINS
                data['cofins_aliquota'] = text_to_float(cofins_match.group(2))  # Alíquota do COFINS
        
        # --- F. ITENS DE FATURA (VALORES DETALHADOS) ---
        # Procura pela seção "Itens de Fatura"
        itens_section = re.search(r'Itens\s*de\s*Fatura.*?(?=ITENS\s*FINANCEIROS|\n\n|\Z)', text_full, re.IGNORECASE | re.DOTALL)
        
        if itens_section:
            itens_text = itens_section.group(0)
            
            # Consumo (kWh) - extrai todos os valores
            consumo_line = re.search(r'Consumo\s*\(kWh\)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,-]+)', itens_text)
            if consumo_line:
                data['consumo_medido'] = text_to_float(consumo_line.group(1))  # Quantidade
                data['preco_unit_consumo'] = text_to_float(consumo_line.group(2))  # Preço Unitário com Tributos
                data['valor_consumo'] = text_to_float(consumo_line.group(6))  # Valor Total
                
                # Se os tributos não foram encontrados na tabela, tenta aqui
                if data['icms'] == 0:
                    data['icms'] = text_to_float(consumo_line.group(5))  # ICMS da linha
            
            # Consumo Compensado (kWh)
            compensado_line = re.search(r'Consumo\s*Compensado[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,-]+)', itens_text)
            if compensado_line:
                data['energia_compensada'] = text_to_float(compensado_line.group(1))  # Quantidade
                data['preco_unit_compensado'] = text_to_float(compensado_line.group(2))  # Preço Unitário com Tributos
                data['valor_consumo_compensado'] = text_to_float(compensado_line.group(6))  # Valor Total
            
            # Energia Injetada
            injetada_line = re.search(r'Energia\s*Inj[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,]+)[^\d]*([\d\.,-]+)', itens_text, re.IGNORECASE)
            if injetada_line:
                data['valor_energia_injetada'] = text_to_float(injetada_line.group(6))  # Valor Total
            
            # Adicional Bandeira
            bandeira_line = re.search(r'Adicional\s*Bandeira[^\d]*([\d\.,-]*)', itens_text, re.IGNORECASE)
            if bandeira_line and bandeira_line.group(1).strip():
                data['valor_adicional_bandeira'] = text_to_float(bandeira_line.group(1))
        
        # --- G. BANDEIRA TARIFÁRIA ---
        bandeira_match = re.search(r'Band\.\s*Tarif\.:\s*([A-Za-z]+)', text_full, re.IGNORECASE)
        if bandeira_match:
            data['bandeira_tarifaria'] = bandeira_match.group(1).strip()
        
        # --- H. INFORMAÇÕES ADICIONAIS ---
        
        # Tipo de Fornecimento
        tipo_match = re.search(r'Tipo\s*de\s*Fornecimento:\s*([A-Z]+)', text_full, re.IGNORECASE)
        if tipo_match:
            data['tipo_fornecimento'] = tipo_match.group(1)
        
        # Classificação
        class_match = re.search(r'Classificação:\s*([A-Za-z]+)', text_full)
        if class_match:
            data['classificacao'] = class_match.group(1)
        
        # CNPJ
        cnpj_match = re.search(r'CNPJ:\s*([\d\.\/-]+)', text_full)
        if cnpj_match:
            data['cnpj'] = cnpj_match.group(1)
        
        # Instalação
        instal_match = re.search(r'INSTALAÇÃO:\s*(\d+)', text_full)
        if instal_match:
            data['instalacao'] = instal_match.group(1)
        
        # --- I. ITENS FINANCEIROS (CIP) ---
        cip_match = re.search(r'Cip[^\d]*([\d\.,]+)', text_full, re.IGNORECASE)
        if not cip_match:
            # Procura na seção de itens financeiros
            financeiros_section = re.search(r'ITENS\s*FINANCEIROS\s*(.*?)(?=\n[A-Z]|\n\n|\Z)', text_full, re.IGNORECASE | re.DOTALL)
            if financeiros_section:
                financeiro_text = financeiros_section.group(1)
                cip_match2 = re.search(r'([\d\.,]+)', financeiro_text)
                if cip_match2:
                    data['valor_cip'] = text_to_float(cip_match2.group(1))
        else:
            data['valor_cip'] = text_to_float(cip_match.group(1))
        
        # --- J. CÁLCULO DO VALOR TOTAL ---
        # Fórmula: Valor Consumo + CIP + Adicional Bandeira - Créditos (Compensado + Injetada)
        data['valor_calculado'] = (
            data['valor_consumo'] + 
            data['valor_cip'] + 
            data['valor_adicional_bandeira'] +
            data['valor_consumo_compensado'] +  # Este geralmente é negativo (crédito)
            data['valor_energia_injetada']      # Este geralmente é negativo (crédito)
        )
        
        # Calcula a diferença entre o valor extraído e o calculado
        if data['total_value'] > 0:
            data['diferenca'] = data['total_value'] - abs(data['valor_calculado'])
        
        # Marca como processado com sucesso
        data['processado'] = True
        
    except Exception as e:
        data['erro'] = str(e)
        print(f"❌ Erro PDF {os.path.basename(pdf_path)}: {e}")
    
    finally:
        if 'doc' in locals():
            doc.close()
    
    return data

# ==========================================
# 3. PROCESSAMENTO EM LOTE
# ==========================================
def processar_faturas(pdf_folder, uc_filtro=None):
    """
    Processa todas as faturas em um diretório
    """
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    
    if not pdf_files:
        print(f"❌ Nenhum PDF encontrado em {pdf_folder}")
        return []
    
    print(f"📡 Processando {len(pdf_files)} faturas...")
    
    resultados = []
    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"  [{i}/{len(pdf_files)}] Processando: {os.path.basename(pdf_path)}")
        dados = extract_invoice_data(pdf_path)
        
        # Adiciona nome do arquivo aos dados
        dados['arquivo'] = os.path.basename(pdf_path)
        
        # Filtra por UC se especificado
        if uc_filtro and dados['uc'] != uc_filtro:
            continue
            
        resultados.append(dados)
    
    print(f"✅ {len(resultados)} faturas processadas com sucesso")
    return resultados

# ==========================================
# 4. GERADOR DE RELATÓRIO COMPLETO
# ==========================================
def gerar_relatorio_completo(resultados, output_path, mes_referencia=None):
    """
    Gera um relatório Excel completo com todos os dados extraídos
    """
    if not resultados:
        print("❌ Nenhum dado para gerar relatório")
        return False
    
    # Converte para DataFrame
    df = pd.DataFrame(resultados)
    
    # Ordena por UC
    if 'uc' in df.columns:
        df = df.sort_values('uc')
    
    # Cria colunas organizadas
    colunas_ordenadas = [
        # IDENTIFICAÇÃO
        'uc', 'instalacao', 'cnpj', 'arquivo',
        
        # DATAS
        'ref_month', 'vencimento', 'data_emissao',
        'dt_anterior', 'dt_atual', 'dt_proxima',
        
        # CONSUMO
        'leitura_ant', 'leitura_atl', 'consumo_medido', 'consumo_faturado',
        'energia_compensada', 'saldo_acumulado',
        
        # VALORES
        'total_value', 'valor_consumo', 'valor_consumo_compensado',
        'valor_energia_injetada', 'valor_cip', 'valor_adicional_bandeira',
        
        # PREÇOS
        'preco_unit_consumo', 'preco_unit_compensado',
        
        # TRIBUTOS - VALORES
        'icms', 'pis', 'cofins',
        
        # TRIBUTOS - ALÍQUOTAS
        'icms_aliquota', 'pis_aliquota', 'cofins_aliquota',
        
        # CÁLCULOS
        'valor_calculado', 'diferenca',
        
        # INFORMAÇÕES ADICIONAIS
        'tipo_fornecimento', 'classificacao', 'bandeira_tarifaria',
        
        # STATUS
        'processado', 'erro'
    ]
    
    # Reorganiza colunas (mantém apenas as que existem no DataFrame)
    colunas_existentes = [col for col in colunas_ordenadas if col in df.columns]
    colunas_restantes = [col for col in df.columns if col not in colunas_existentes]
    
    df = df[colunas_existentes + colunas_restantes]
    
    # Salva para Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Faturas Detalhadas', index=False)
        
        # Adiciona sumário
        ws_summary = writer.book.create_sheet(title='Sumário')
        
        # Estatísticas
        total_faturas = len(df)
        faturas_processadas = df['processado'].sum() if 'processado' in df.columns else 0
        total_valor = df['total_value'].sum() if 'total_value' in df.columns else 0
        total_icms = df['icms'].sum() if 'icms' in df.columns else 0
        
        ws_summary.append(['RELATÓRIO DE FATURAS - EQUATORIAL'])
        ws_summary.append([''])
        ws_summary.append([f'Mês de Referência: {mes_referencia or "Não especificado"}' ])
        ws_summary.append([f'Data de Geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
        ws_summary.append([''])
        ws_summary.append(['ESTATÍSTICAS:'])
        ws_summary.append([f'Total de Faturas: {total_faturas}'])
        ws_summary.append([f'Faturas Processadas: {faturas_processadas}'])
        ws_summary.append([f'Faturas com Erro: {total_faturas - faturas_processadas}'])
        ws_summary.append([f'Valor Total: R$ {total_valor:,.2f}'])
        ws_summary.append([f'ICMS Total: R$ {total_icms:,.2f}'])
        
        # Formata largura das colunas
        for sheet in writer.book.sheetnames:
            ws = writer.book[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Relatório salvo em: {output_path}")
    return True

# ==========================================
# 5. FUNÇÃO PARA PROCESSAR CLIENTE ESPECÍFICO
# ==========================================
def processar_cliente_especifico(pdf_folder, uc_cliente, output_dir=None):
    """
    Processa faturas de um cliente específico (por UC)
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(pdf_folder), "cliente_especifico")
    
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"🔍 Procurando faturas para UC: {uc_cliente}")
    
    resultados = processar_faturas(pdf_folder, uc_filtro=uc_cliente)
    
    if not resultados:
        print(f"❌ Nenhuma fatura encontrada para UC {uc_cliente}")
        return None
    
    # Salva dados em JSON
    json_path = os.path.join(output_dir, f"dados_uc_{uc_cliente}.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2, default=str)
    
    # Gera relatório Excel
    excel_path = os.path.join(output_dir, f"relatorio_uc_{uc_cliente}.xlsx")
    gerar_relatorio_completo(resultados, excel_path, f"UC {uc_cliente}")
    
    print(f"✅ Dados do cliente {uc_cliente} salvos em:")
    print(f"   JSON: {json_path}")
    print(f"   Excel: {excel_path}")
    
    return resultados

# ==========================================
# 6. FUNÇÃO PARA GERAR RELATÓRIO GERAL
# ==========================================
def gerar_relatorio_geral(pdf_folder, mes_referencia, output_dir=None):
    """
    Gera relatório geral de todas as faturas
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(pdf_folder), "relatorios")
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Processa todas as faturas
    resultados = processar_faturas(pdf_folder)
    
    if not resultados:
        print("❌ Nenhuma fatura processada")
        return False
    
    # Gera nome do arquivo
    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"Relatorio_Faturas_{mes_referencia.replace('/', '-')}_{data_hora}.xlsx"
    output_path = os.path.join(output_dir, nome_arquivo)
    
    # Gera relatório
    sucesso = gerar_relatorio_completo(resultados, output_path, mes_referencia)
    
    # Gera também um resumo consolidado
    if sucesso:
        gerar_resumo_consolidado(resultados, output_dir, mes_referencia)
    
    return sucesso

def gerar_resumo_consolidado(resultados, output_dir, mes_referencia):
    """
    Gera um resumo consolidado simplificado
    """
    df = pd.DataFrame(resultados)
    
    # Colunas para o resumo
    colunas_resumo = [
        'uc', 'ref_month', 'vencimento', 'total_value',
        'consumo_medido', 'energia_compensada', 'icms',
        'pis', 'cofins', 'valor_cip', 'processado'
    ]
    
    # Filtra colunas existentes
    colunas_existentes = [col for col in colunas_resumo if col in df.columns]
    df_resumo = df[colunas_existentes].copy()
    
    # Adiciona status
    if 'processado' in df_resumo.columns:
        df_resumo['status'] = df_resumo['processado'].apply(lambda x: '✅ OK' if x else '❌ Erro')
    
    # Salva resumo
    resumo_path = os.path.join(output_dir, f"Resumo_{mes_referencia.replace('/', '-')}.xlsx")
    df_resumo.to_excel(resumo_path, index=False)
    
    print(f"✅ Resumo consolidado salvo em: {resumo_path}")
    
    return True